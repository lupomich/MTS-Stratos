#!/usr/bin/env python3
"""
Generate Excel report from test-results.json
"""

import json
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def load_results(json_path):
    """Load test results from JSON file"""
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"ERROR: {json_path} not found")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"ERROR: Invalid JSON: {e}")
        sys.exit(1)

def create_excel_report(data, output_path):
    """Create Excel workbook with Summary and Details sheets"""
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # === SUMMARY SHEET ===
    ws_summary = wb.create_sheet('Summary', 0)
    
    # Title
    ws_summary['A1'] = 'MTS-Stratos E2E Test Report - Summary'
    ws_summary['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws_summary['A1'].fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
    ws_summary.merge_cells('A1:D1')
    
    # Execution info
    summary = data.get('summary', {})
    row = 3
    info_data = [
        ('Execution Date', summary.get('startTime', 'N/A')),
        ('End Time', summary.get('endTime', 'N/A')),
        ('Total Duration', f"{summary.get('durationMs', 0) / 1000:.2f} seconds"),
        ('', ''),
        ('Total Tests', summary.get('totalTests', 0)),
        ('PASSED', summary.get('passed', 0)),
        ('FAILED', summary.get('failed', 0)),
        ('Pass Rate', summary.get('passRate', '0%')),
    ]
    
    for label, value in info_data:
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = value
        if label in ['PASSED', 'FAILED']:
            ws_summary[f'A{row}'].font = Font(bold=True)
            ws_summary[f'B{row}'].font = Font(bold=True)
            if label == 'PASSED':
                ws_summary[f'B{row}'].fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
            else:
                ws_summary[f'B{row}'].fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
        row += 1
    
    # Section breakdown
    row += 2
    ws_summary[f'A{row}'] = 'Section Breakdown'
    ws_summary[f'A{row}'].font = Font(size=14, bold=True)
    row += 1
    
    # Headers
    headers = ['Section', 'Total Tests', 'Passed', 'Failed', 'Pass Rate']
    for col, header in enumerate(headers, start=1):
        cell = ws_summary.cell(row, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    row += 1
    
    # Calculate section stats
    tests = data.get('tests', [])
    sections = {
        'Section 1: User Management (T01-T24)': [t for t in tests if t['id'].startswith('T0') or t['id'] in [f'T{i}' for i in range(10, 25)]],
        'Section 2: Settings Persistence (T25-T36)': [t for t in tests if t['id'] in [f'T{i}' for i in range(25, 37)]],
        'Section 3: Integration & Cleanup (T37-T40)': [t for t in tests if t['id'] in [f'T{i}' for i in range(37, 41)]],
    }
    
    for section_name, section_tests in sections.items():
        total = len(section_tests)
        passed = sum(1 for t in section_tests if t['status'] == 'PASS')
        failed = total - passed
        pass_rate = f"{(passed/total*100):.1f}%" if total > 0 else "0%"
        
        ws_summary.cell(row, 1, section_name)
        ws_summary.cell(row, 2, total)
        ws_summary.cell(row, 3, passed)
        ws_summary.cell(row, 4, failed)
        ws_summary.cell(row, 5, pass_rate)
        
        # Color code pass rate
        if passed == total:
            ws_summary.cell(row, 5).fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
        elif failed > 0:
            ws_summary.cell(row, 5).fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
        
        row += 1
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 12
    ws_summary.column_dimensions['D'].width = 12
    ws_summary.column_dimensions['E'].width = 12
    
    # === DETAILS SHEET ===
    ws_details = wb.create_sheet('Test Details', 1)
    
    # Title
    ws_details['A1'] = 'Test Execution Details'
    ws_details['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws_details['A1'].fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
    ws_details.merge_cells('A1:G1')
    
    # Headers
    row = 3
    headers = ['Test ID', 'Description', 'Type', 'Start Time', 'Duration (ms)', 'Status', 'Fail Reason']
    for col, header in enumerate(headers, start=1):
        cell = ws_details.cell(row, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    row += 1
    
    # Test rows
    for test in tests:
        ws_details.cell(row, 1, test.get('id', ''))
        ws_details.cell(row, 2, test.get('description', ''))
        ws_details.cell(row, 3, test.get('type', ''))
        ws_details.cell(row, 4, test.get('startTime', ''))
        ws_details.cell(row, 5, test.get('duration', 0))
        ws_details.cell(row, 6, test.get('status', ''))
        ws_details.cell(row, 7, test.get('failReason', '-'))
        
        # Color code status
        status_cell = ws_details.cell(row, 6)
        if test.get('status') == 'PASS':
            status_cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
            status_cell.font = Font(bold=True, color='155724')
        elif test.get('status') == 'FAIL':
            status_cell.fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
            status_cell.font = Font(bold=True, color='721C24')
        
        # Add borders
        for col in range(1, 8):
            ws_details.cell(row, col).border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        row += 1
    
    # Adjust column widths
    ws_details.column_dimensions['A'].width = 10
    ws_details.column_dimensions['B'].width = 35
    ws_details.column_dimensions['C'].width = 8
    ws_details.column_dimensions['D'].width = 20
    ws_details.column_dimensions['E'].width = 15
    ws_details.column_dimensions['F'].width = 10
    ws_details.column_dimensions['G'].width = 50
    
    # Freeze panes
    ws_details.freeze_panes = 'A4'
    
    # Save workbook
    wb.save(output_path)
    print(f"✅ Excel report saved: {output_path}")

def main():
    json_path = 'test-results.json'
    output_path = 'TEST_RESULTS.xlsx'
    
    if len(sys.argv) > 1:
        json_path = sys.argv[1]
    if len(sys.argv) > 2:
        output_path = sys.argv[2]
    
    print(f"Loading test results from: {json_path}")
    data = load_results(json_path)
    
    print(f"Generating Excel report: {output_path}")
    create_excel_report(data, output_path)
    
    print("Done!")

if __name__ == '__main__':
    main()
