#!/usr/bin/env python3
"""
Generate Excel file from test checklist
Requires: openpyxl
Install: pip install openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Test Checklist"

# Define styles
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
section_font = Font(bold=True, size=11)
checkbox_font = Font(name="Courier New", size=10)
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Set column widths
ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 50
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 40

# Headers
headers = ['Test #', 'Description', 'Status', 'Pass/Fail', 'Notes']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Test data
tests = [
    # Section 1: Admin User
    ('1', '1.1', 'Create user Admin (via Admin Panel)', '☐', ''),
    ('1', '1.2', 'Login user Admin-test', '☐', ''),
    ('1', '1.3', 'Logout user Admin-test', '☐', ''),
    ('1', '1.4', 'Disable user Admin-test', '☐', ''),
    ('1', '1.5', 'Try login disabled Admin-test (should fail)', '☐', ''),
    ('1', '1.6', 'Re-enable user Admin-test', '☐', ''),
    ('1', '1.7', 'Login after re-enable Admin-test', '☐', ''),
    ('1', '1.8', 'Delete user Admin-test', '☐', ''),

    # Section 2: Member User
    ('2', '2.1', 'Create user Member', '☐', ''),
    ('2', '2.2', 'Login user Member-test', '☐', ''),
    ('2', '2.3', 'Logout user Member-test', '☐', ''),
    ('2', '2.4', 'Disable user Member-test', '☐', ''),
    ('2', '2.5', 'Try login disabled Member-test (should fail)', '☐', ''),
    ('2', '2.6', 'Re-enable user Member-test', '☐', ''),
    ('2', '2.7', 'Login after re-enable Member-test', '☐', ''),
    ('2', '2.8', 'Delete user Member-test', '☐', ''),

    # Section 3: Trader User
    ('3', '3.1', 'Create user Trader', '☐', ''),
    ('3', '3.2', 'Login user Trader-test', '☐', ''),
    ('3', '3.3', 'Logout user Trader-test', '☐', ''),
    ('3', '3.4', 'Disable user Trader-test', '☐', ''),
    ('3', '3.5', 'Try login disabled Trader-test (should fail)', '☐', ''),
    ('3', '3.6', 'Re-enable user Trader-test', '☐', ''),
    ('3', '3.7', 'Login after re-enable Trader-test', '☐', ''),
    ('3', '3.8', 'Delete user Trader-test', '☐', ''),

    # Section 4: AutoEx User
    ('4', '4.1', 'Create user AutoEx', '☐', ''),
    ('4', '4.2', 'Login user AutoEx-test', '☐', ''),
    ('4', '4.3', 'Logout user AutoEx-test', '☐', ''),
    ('4', '4.4', 'Disable user AutoEx-test', '☐', ''),
    ('4', '4.5', 'Try login disabled AutoEx-test (should fail)', '☐', ''),
    ('4', '4.6', 'Re-enable user AutoEx-test', '☐', ''),
    ('4', '4.7', 'Login after re-enable AutoEx-test', '☐', ''),
    ('4', '4.8', 'Delete user AutoEx-test', '☐', ''),

    # Section 5: UI Settings
    ('5', '5.1', 'Login and view columns (Member-test)', '☐', ''),
    ('5', '5.2', 'Hide column (PRICE)', '☐', ''),
    ('5', '5.3', 'Drag & drop column reorder', '☐', ''),
    ('5', '5.4', 'Apply single filter', '☐', ''),
    ('5', '5.5', 'Apply multiple filters', '☐', ''),
    ('5', '5.6', 'Logout and verify settings persist', '☐', ''),
    ('5', '5.7', 'Reset all columns', '☐', ''),
    ('5', '5.8', 'Verify reset persists after logout', '☐', ''),

    # Section 6: UI Features
    ('6', '6.1', 'Verify header status badges (TEST, Market, Member, etc.)', '☐', ''),
    ('6', '6.2', 'Test responsive UI (mobile resize)', '☐', ''),
    ('6', '6.3', 'Change language (EN ↔ IT)', '☐', ''),
    ('6', '6.4', 'Change theme (dark/light)', '☐', ''),

    # Section 7: Database
    ('7', '7.1', 'Verify database has no test users post-test', '☐', ''),
    ('7', '7.2', 'Verify cache/session cleanup', '☐', ''),

    # Section 8: Edge Cases
    ('8', '8.1', 'Simultaneous login same user (2 browsers)', '☐', ''),
    ('8', '8.2', 'Session timeout handling', '☐', ''),
    ('8', '8.3', 'Change password and verify old fails', '☐', ''),
    ('8', '8.4', 'Browser crash recovery (UI state)', '☐', ''),
]

# Add test rows
row = 2
for section, test_id, description, checkbox, notes in tests:
    ws.cell(row=row, column=1, value=test_id)
    ws.cell(row=row, column=2, value=description)
    ws.cell(row=row, column=3, value=checkbox)
    ws.cell(row=row, column=4, value='')
    ws.cell(row=row, column=5, value=notes)

    for col in range(1, 6):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    row += 1

# Add summary section
row += 2
ws.cell(row=row, column=1, value='SUMMARY').font = section_font
ws.cell(row=row, column=1).fill = section_fill

row += 1
summary_rows = [
    ('Start Time:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
    ('End Time:', ''),
    ('Total Duration:', ''),
    ('Tests Passed:', ''),
    ('Tests Failed:', ''),
    ('Success Rate:', ''),
    ('Tester Name:', ''),
    ('Environment:', '☐ Dev  ☐ Staging  ☐ Prod'),
    ('Blocker Issues:', ''),
    ('Notes:', ''),
]

for label, value in summary_rows:
    ws.cell(row=row, column=1, value=label).font = Font(bold=True)
    ws.cell(row=row, column=2, value=value)
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    row += 1

# Create a second sheet for detailed results
ws_results = wb.create_sheet("Test Results")

# Results sheet headers
result_headers = ['Test #', 'Description', 'Result', 'Start Time', 'End Time', 'Duration (s)', 'Error Details']
for col, header in enumerate(result_headers, 1):
    cell = ws_results.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

ws_results.column_dimensions['A'].width = 8
ws_results.column_dimensions['B'].width = 50
ws_results.column_dimensions['C'].width = 12
ws_results.column_dimensions['D'].width = 20
ws_results.column_dimensions['E'].width = 20
ws_results.column_dimensions['F'].width = 12
ws_results.column_dimensions['G'].width = 40

# Save workbook
output_path = 'TEST_CHECKLIST.xlsx'
wb.save(output_path)
print(f"✅ Excel file created: {output_path}")
